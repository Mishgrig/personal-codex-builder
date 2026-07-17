from __future__ import annotations

import base64
import csv
import io
import json
import re
from collections.abc import Iterable
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select, text
from sqlalchemy.orm import Session, selectinload

from app.core.exceptions import NotFoundError, ValidationAPIError
from app.schemas.card import CardCreate, CardUpdate
from app.models.workspace import (
    Asset,
    CardSchema,
    CardRegistry,
    CardTypeDefinition,
    CardTypeField,
    SchemaFieldDefinition,
)
from app.schemas.schema import CardSchemaCreate
from app.services.search_service import rebuild_index


def _safe_slug(value: str, *, fallback: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", value.strip().lower())
    normalized = normalized.strip("-")
    return normalized or fallback


def card_type_table_name(card_type_slug: str) -> str:
    return f"card_type_{_safe_slug(card_type_slug, fallback='generic')}"


def card_type_column_name(field_slug: str) -> str:
    return f"field_{_safe_slug(field_slug, fallback='value').replace('-', '_')}"


def _sqlite_type(kind: str) -> str:
    if kind in {"number", "rating"}:
        return "REAL"
    if kind in {"boolean"}:
        return "INTEGER"
    return "TEXT"


def _escape_identifier(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _upsert_legacy_schema(session: Session, payload: CardSchemaCreate) -> CardSchema:
    schema = session.get(CardSchema, payload.id)
    if schema is None:
        schema = CardSchema(id=payload.id)
    schema.label = payload.label
    schema.description = payload.description
    schema.icon = payload.icon
    schema.field_order = payload.field_order or [field.field_id for field in payload.fields]
    schema.is_active = payload.is_active

    existing_fields = {field.field_id: field for field in schema.fields}
    incoming_ids = {field.field_id for field in payload.fields}
    for field_id, existing in list(existing_fields.items()):
        if field_id not in incoming_ids:
            schema.fields.remove(existing)

    for item in payload.fields:
        field = existing_fields.get(item.field_id)
        if field is None:
            field = SchemaFieldDefinition(field_id=item.field_id)
            schema.fields.append(field)
        field.label = item.label
        field.kind = item.kind
        field.description = item.description
        field.required = item.required
        field.repeatable = item.repeatable
        field.default_value = item.default_value
        field.options = item.options
        field.placeholder = item.placeholder
        field.show_in_card = item.show_in_card
        field.show_in_list = item.show_in_list
        field.show_in_filters = item.show_in_filters
        field.validation = item.validation
        field.sort_order = item.sort_order
        field.is_active = item.is_active
    session.add(schema)
    return schema


def _ensure_card_type_table(session: Session, card_type: CardTypeDefinition, fields: Iterable[CardTypeField]) -> None:
    table_name = _escape_identifier(card_type.table_name)
    session.execute(
        text(
            f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                card_id INTEGER PRIMARY KEY,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP NOT NULL,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP NOT NULL
            )
            """
        )
    )
    existing_columns = {
        row[1]
        for row in session.execute(text(f"PRAGMA table_info({_escape_identifier(card_type.table_name)})")).all()
    }
    for field in fields:
        if field.sql_column_name in existing_columns:
            continue
        session.execute(
            text(
                f"ALTER TABLE {table_name} ADD COLUMN {_escape_identifier(field.sql_column_name)} {_sqlite_type(field.field_type)}"
            )
        )


def _upsert_card_type(session: Session, payload: CardSchemaCreate) -> CardTypeDefinition:
    slug = _safe_slug(payload.id, fallback="general")
    table_name = card_type_table_name(slug)
    definition = session.get(CardTypeDefinition, slug)
    if definition is None:
        definition = CardTypeDefinition(slug=slug, table_name=table_name)
    definition.name = payload.label
    definition.description = payload.description
    definition.icon = payload.icon
    definition.is_active = payload.is_active
    definition.layout_json = {
        **(payload.layout_json or {}),
        "field_order": payload.field_order or [field.field_id for field in payload.fields],
        "mode": "card-type-studio",
    }
    session.add(definition)
    session.flush()

    existing_fields = {field.field_slug: field for field in definition.fields}
    incoming_slugs = {_safe_slug(field.field_id, fallback=f"field-{index}") for index, field in enumerate(payload.fields)}
    for field_slug, existing in list(existing_fields.items()):
        if field_slug not in incoming_slugs:
            existing.is_active = False

    created_fields: list[CardTypeField] = []
    for index, item in enumerate(payload.fields):
        field_slug = _safe_slug(item.field_id, fallback=f"field-{index}")
        sql_column_name = card_type_column_name(field_slug)
        field = existing_fields.get(field_slug)
        if field is None:
            field = CardTypeField(card_type_slug=definition.slug, field_slug=field_slug, sql_column_name=sql_column_name)
            definition.fields.append(field)
        field.name = item.label
        field.field_type = item.kind
        field.required = item.required
        field.visible = item.show_in_card or item.show_in_list
        field.show_in_card = item.show_in_card
        field.show_in_atlas = item.show_in_list
        field.include_in_table_view = True
        field.include_in_export = True
        field.allow_import = True
        field.searchable = item.kind in {"text", "long_text", "markdown", "url", "select", "multi_select"}
        field.filterable = item.show_in_filters
        field.description = item.description
        field.help_text = item.placeholder
        field.default_value_json = item.default_value
        field.options_json = item.options
        field.sort_order = index
        field.is_active = item.is_active
        created_fields.append(field)

    session.flush()
    _ensure_card_type_table(session, definition, created_fields)
    return definition


def _schema_query():
    return (
        select(CardSchema)
        .options(selectinload(CardSchema.fields))
        .order_by(CardSchema.label.asc())
    )


def list_schemas(session: Session) -> list[CardSchema]:
    schemas = list(session.execute(_schema_query()).scalars())
    definitions = {
        definition.slug: definition
        for definition in session.execute(select(CardTypeDefinition)).scalars().all()
    }
    for schema in schemas:
        definition = definitions.get(_safe_slug(schema.id, fallback="general"))
        schema.layout_json = definition.layout_json if definition else {}
    return schemas


def upsert_schema(session: Session, payload: CardSchemaCreate) -> CardSchema:
    _upsert_legacy_schema(session, payload)
    definition = _upsert_card_type(session, payload)
    session.commit()
    schema = session.execute(_schema_query().where(CardSchema.id == payload.id)).scalar_one()
    schema.layout_json = definition.layout_json or {}
    return schema


def _card_type_query():
    return (
        select(CardTypeDefinition)
        .options(selectinload(CardTypeDefinition.fields))
        .order_by(CardTypeDefinition.name.asc())
    )


def _get_card_type_or_raise(session: Session, card_type_slug: str) -> CardTypeDefinition:
    definition = session.execute(
        _card_type_query().where(CardTypeDefinition.slug == _safe_slug(card_type_slug, fallback="general"))
    ).scalar_one_or_none()
    if definition is None:
        raise NotFoundError("Card type was not found.", code="CARD_TYPE_NOT_FOUND")
    return definition


def list_card_types(session: Session) -> list[dict[str, Any]]:
    definitions = list(session.execute(_card_type_query()).scalars())
    counts = dict(
        session.execute(
            select(CardRegistry.card_type_slug, func.count(CardRegistry.id))
            .where(CardRegistry.deleted_at.is_(None))
            .group_by(CardRegistry.card_type_slug)
        ).all()
    )
    items: list[dict[str, Any]] = []
    for definition in definitions:
        items.append(
            {
                "slug": definition.slug,
                "name": definition.name,
                "table_name": definition.table_name,
                "description": definition.description,
                "icon": definition.icon,
                "is_active": definition.is_active,
                "layout_json": definition.layout_json or {},
                "fields": [field for field in definition.fields if field.is_active],
                "card_count": counts.get(definition.slug, 0),
                "created_at": definition.created_at,
                "updated_at": definition.updated_at,
            }
        )
    return items


def get_card_type_table(session: Session, card_type_slug: str, *, q: str = "") -> dict[str, Any]:
    return get_card_type_table_with_filters(session, card_type_slug, q=q)


def get_card_type_table_with_filters(
    session: Session,
    card_type_slug: str,
    *,
    q: str = "",
    sort_by: str = "manual",
    sort_dir: str = "asc",
    status: str = "",
) -> dict[str, Any]:
    definition = _get_card_type_or_raise(session, card_type_slug)
    fields = [field for field in definition.fields if field.is_active and field.include_in_table_view]

    select_columns = [
        "cards.id AS card_id",
        "cards_registry.id AS registry_id",
        "cards.title AS title",
        "cards.summary AS summary",
        "cards.status AS status",
    ]
    for field in fields:
        select_columns.append(
            f"{_escape_identifier(definition.table_name)}.{_escape_identifier(field.sql_column_name)} AS {_escape_identifier(field.sql_column_name)}"
        )

    where_sql = "WHERE cards_registry.card_type_slug = :card_type_slug AND cards_registry.deleted_at IS NULL"
    params: dict[str, Any] = {"card_type_slug": definition.slug}
    if status.strip():
        params["status"] = status.strip()
        where_sql += " AND cards.status = :status"
    if q.strip():
        params["q"] = f"%{q.strip().lower()}%"
        where_sql += (
            " AND (lower(cards.title) LIKE :q OR lower(cards.summary) LIKE :q OR lower(cards.body_text) LIKE :q"
        )
        for index, field in enumerate(fields):
            param_name = f"field_q_{index}"
            params[param_name] = f"%{q.strip().lower()}%"
            where_sql += (
                f" OR lower(COALESCE(CAST({_escape_identifier(definition.table_name)}.{_escape_identifier(field.sql_column_name)} AS TEXT), '')) LIKE :{param_name}"
            )
        where_sql += ")"

    order_sql = "cards.sort_order ASC, cards.id ASC"
    if sort_by == "title":
        order_sql = f"cards.title {'DESC' if sort_dir == 'desc' else 'ASC'}"
    elif sort_by == "status":
        order_sql = f"cards.status {'DESC' if sort_dir == 'desc' else 'ASC'}"
    elif sort_by == "summary":
        order_sql = f"cards.summary {'DESC' if sort_dir == 'desc' else 'ASC'}"
    else:
        selected_field = next((field for field in fields if field.field_slug == sort_by), None)
        if selected_field is not None:
            order_sql = (
                f"{_escape_identifier(definition.table_name)}.{_escape_identifier(selected_field.sql_column_name)} "
                f"{'DESC' if sort_dir == 'desc' else 'ASC'}, cards.sort_order ASC, cards.id ASC"
            )

    query = text(
        f"""
        SELECT {", ".join(select_columns)}
        FROM cards_registry
        JOIN cards ON cards.id = cards_registry.legacy_card_id
        LEFT JOIN {_escape_identifier(definition.table_name)}
            ON {_escape_identifier(definition.table_name)}.card_id = cards.id
        {where_sql}
        ORDER BY {order_sql}
        """
    )
    rows = session.execute(query, params).mappings().all()

    column_items = [
        {
            "field_slug": field.field_slug,
            "sql_column_name": field.sql_column_name,
            "name": field.name,
            "field_type": field.field_type,
            "required": field.required,
            "searchable": field.searchable,
            "filterable": field.filterable,
        }
        for field in fields
    ]
    row_items = [
        {
            "card_id": row["card_id"],
            "registry_id": row["registry_id"],
            "title": row["title"],
            "summary": row["summary"],
            "status": row["status"],
            "values": {field.field_slug: row.get(field.sql_column_name) for field in fields},
        }
        for row in rows
    ]
    definition_payload = next(item for item in list_card_types(session) if item["slug"] == definition.slug)
    return {
        "card_type": definition_payload,
        "columns": column_items,
        "rows": row_items,
        "total": len(row_items),
        "q": q,
        "sort_by": sort_by,
        "sort_dir": sort_dir,
        "status": status,
    }


def create_card_type_row(session: Session, card_type_slug: str, *, title: str, summary: str, status: str, values: dict[str, Any]) -> dict[str, Any]:
    from app.schemas.card import CardUpdate
    from app.services.card_service import create_card, update_card

    definition = _get_card_type_or_raise(session, card_type_slug)
    created_card = create_card(
        session,
        CardCreate(
            title=title or "Untitled Card",
            summary=summary,
            status=status or "draft",
            schema_id=definition.slug,
            taxonomy_term_ids=[],
        ),
    )
    updated_card = update_card(
        session,
        created_card.id,
        CardUpdate(
            title=title or created_card.title,
            summary=summary,
            status=status or "draft",
            schema_id=definition.slug,
            dynamic_fields=values,
        ),
    )
    row = next((item for item in get_card_type_table_with_filters(session, definition.slug)["rows"] if item["card_id"] == updated_card.id), None)
    return row or {
        "card_id": updated_card.id,
        "registry_id": updated_card.id,
        "title": updated_card.title,
        "summary": updated_card.summary,
        "status": updated_card.status,
        "values": updated_card.dynamic_fields,
    }


def update_card_type_row(
    session: Session,
    card_type_slug: str,
    card_id: int,
    *,
    title: str | None,
    summary: str | None,
    status: str | None,
    values: dict[str, Any],
) -> dict[str, Any]:
    from app.schemas.card import CardUpdate
    from app.services.card_service import get_card, update_card

    definition = _get_card_type_or_raise(session, card_type_slug)
    existing = get_card(session, card_id)
    if (existing.schema_id or "general") != definition.slug:
        raise NotFoundError("Card type row was not found.", code="CARD_TYPE_ROW_NOT_FOUND")
    updated_card = update_card(
        session,
        card_id,
        CardUpdate(
            title=title,
            summary=summary,
            status=status,
            schema_id=definition.slug,
            dynamic_fields=values,
        ),
    )
    row = next((item for item in get_card_type_table_with_filters(session, definition.slug)["rows"] if item["card_id"] == updated_card.id), None)
    return row or {
        "card_id": updated_card.id,
        "registry_id": updated_card.id,
        "title": updated_card.title,
        "summary": updated_card.summary,
        "status": updated_card.status,
        "values": updated_card.dynamic_fields,
    }


def soft_delete_card_type_row(session: Session, card_type_slug: str, card_id: int) -> dict[str, Any]:
    from app.services.card_service import delete_card, get_card

    definition = _get_card_type_or_raise(session, card_type_slug)
    existing = get_card(session, card_id)
    if (existing.schema_id or "general") != definition.slug:
        raise NotFoundError("Card type row was not found.", code="CARD_TYPE_ROW_NOT_FOUND")
    delete_card(session, card_id)
    return {"card_id": card_id, "deleted": True}


def export_card_type_structure(session: Session, card_type_slug: str, *, export_format: str) -> dict[str, Any]:
    definition = _get_card_type_or_raise(session, card_type_slug)
    fields = [field for field in definition.fields if field.is_active]
    export_rows = [
        {
            "card_type_name": definition.name,
            "card_type_slug": definition.slug,
            "sql_table_name": definition.table_name,
            "field_name": field.name,
            "field_slug": field.field_slug,
            "sql_column_name": field.sql_column_name,
            "field_type": field.field_type,
            "required": field.required,
            "searchable": field.searchable,
            "importable": field.allow_import,
            "exportable": field.include_in_export,
            "description": field.description,
            "help_text": field.help_text,
        }
        for field in fields
    ]
    export_format = export_format.lower()
    if export_format == "json":
        return {
            "card_type_slug": definition.slug,
            "format": "json",
            "filename": f"{definition.slug}-structure.json",
            "content_json": export_rows,
            "content_text": "",
        }
    if export_format == "csv":
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=list(export_rows[0].keys()) if export_rows else [
            "card_type_name",
            "card_type_slug",
            "sql_table_name",
            "field_name",
            "field_slug",
            "sql_column_name",
            "field_type",
            "required",
            "searchable",
            "importable",
            "exportable",
            "description",
            "help_text",
        ])
        writer.writeheader()
        writer.writerows(export_rows)
        return {
            "card_type_slug": definition.slug,
            "format": "csv",
            "filename": f"{definition.slug}-structure.csv",
            "content_json": [],
            "content_text": buffer.getvalue(),
            "content_base64": "",
        }
    if export_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Card Type Structure"
        headers = list(export_rows[0].keys()) if export_rows else [
            "card_type_name",
            "card_type_slug",
            "sql_table_name",
            "field_name",
            "field_slug",
            "sql_column_name",
            "field_type",
            "required",
            "searchable",
            "importable",
            "exportable",
            "description",
            "help_text",
        ]
        sheet.append(headers)
        for row in export_rows:
            sheet.append([row.get(header) for header in headers])
        buffer = io.BytesIO()
        workbook.save(buffer)
        return {
            "card_type_slug": definition.slug,
            "format": "xlsx",
            "filename": f"{definition.slug}-structure.xlsx",
            "content_json": [],
            "content_text": "",
            "content_base64": base64.b64encode(buffer.getvalue()).decode("ascii"),
        }
    raise ValidationAPIError("Unsupported export format.", details={"supported": ["json", "csv", "xlsx"]})


def _parse_import_payload(
    *,
    content_text: str,
    content_base64: str,
    import_format: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    import_format = import_format.lower()
    if import_format == "json":
        parsed = json.loads(content_text or "[]")
        if isinstance(parsed, dict):
            rows = parsed.get("rows", [])
            asset_files = parsed.get("asset_files", [])
            if not isinstance(rows, list) or any(not isinstance(row, dict) for row in rows):
                raise ValidationAPIError("JSON bundle import must contain a rows list.")
            if not isinstance(asset_files, list) or any(not isinstance(item, dict) for item in asset_files):
                raise ValidationAPIError("JSON bundle import asset_files must be a list of objects.")
            return rows, asset_files
        if not isinstance(parsed, list) or any(not isinstance(row, dict) for row in parsed):
            raise ValidationAPIError("JSON import must be a list of objects.")
        return parsed, []
    if import_format == "csv":
        buffer = io.StringIO(content_text)
        reader = csv.DictReader(buffer)
        return [dict(row) for row in reader], []
    if import_format == "xlsx":
        if not content_base64:
            raise ValidationAPIError("XLSX import requires binary workbook content.")
        workbook = load_workbook(io.BytesIO(base64.b64decode(content_base64)))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        parsed_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            parsed_rows.append(
                {
                    headers[index]: row[index]
                    for index in range(min(len(headers), len(row)))
                    if headers[index]
                }
            )
        return parsed_rows, []
    raise ValidationAPIError("Unsupported import format.", details={"supported": ["json", "csv", "xlsx"]})


def _import_mapping(definition: CardTypeDefinition) -> dict[str, str]:
    mapping = {
        "card_id": "card_id",
        "title": "title",
        "summary": "summary",
        "status": "status",
        "cover_asset_id": "cover_asset_id",
        "gallery_asset_ids": "gallery_asset_ids",
        "attachment_asset_ids": "attachment_asset_ids",
    }
    for field in definition.fields:
        if not field.is_active or not field.allow_import:
            continue
        mapping[field.field_slug] = field.field_slug
        mapping[field.name.strip().lower()] = field.field_slug
        mapping[field.sql_column_name] = field.field_slug
    return mapping


def preview_card_type_import(
    session: Session,
    card_type_slug: str,
    *,
    import_format: str,
    content_text: str,
    content_base64: str = "",
) -> dict[str, Any]:
    definition = _get_card_type_or_raise(session, card_type_slug)
    rows, _asset_files = _parse_import_payload(
        content_text=content_text,
        content_base64=content_base64,
        import_format=import_format,
    )
    mapping = _import_mapping(definition)
    headers = list(rows[0].keys()) if rows else []
    normalized_headers = {header: mapping.get(header.strip().lower(), mapping.get(header, "")) for header in headers}
    missing_columns = [
        field.field_slug
        for field in definition.fields
        if field.is_active and field.allow_import and field.required and field.field_slug not in normalized_headers.values()
    ]
    unknown_columns = [header for header, field_slug in normalized_headers.items() if not field_slug]
    matched_columns = {header: field_slug for header, field_slug in normalized_headers.items() if field_slug}
    return {
        "card_type_slug": definition.slug,
        "format": import_format.lower(),
        "row_count": len(rows),
        "missing_columns": missing_columns,
        "unknown_columns": unknown_columns,
        "matched_columns": matched_columns,
        "sample_rows": rows[:5],
    }


def _coerce_import_value(field: CardTypeField, value: Any) -> Any:
    if value is None or value == "":
        return None
    if field.field_type in {"number", "rating"}:
        return float(value)
    if field.field_type == "boolean":
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "y"}
    if field.field_type == "multi_select":
        if isinstance(value, list):
            return value
        return [item.strip() for item in str(value).split(",") if item.strip()]
    return value


def apply_card_type_import(
    session: Session,
    card_type_slug: str,
    *,
    workspace_slug: str,
    import_format: str,
    content_text: str,
    content_base64: str = "",
) -> dict[str, Any]:
    from app.services.card_service import create_card, get_card, update_card
    from app.services.file_service import attach_existing_asset, register_imported_asset

    definition = _get_card_type_or_raise(session, card_type_slug)
    preview = preview_card_type_import(
        session,
        card_type_slug,
        import_format=import_format,
        content_text=content_text,
        content_base64=content_base64,
    )
    rows, asset_files = _parse_import_payload(
        content_text=content_text,
        content_base64=content_base64,
        import_format=import_format,
    )
    fields = {
        field.field_slug: field
        for field in definition.fields
        if field.is_active and field.allow_import
    }
    matched_columns = preview["matched_columns"]
    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    imported_assets: dict[str, str] = {}

    for asset_file in asset_files:
        try:
            asset_id = str(asset_file.get("asset_id") or "").strip() or None
            original_filename = str(asset_file.get("filename") or "").strip()
            content_b64 = str(asset_file.get("content_base64") or "").strip()
            mime_type = str(asset_file.get("mime_type") or "").strip()
            if not original_filename or not content_b64:
                raise ValidationAPIError("Bundled asset files require filename and content_base64.")
            imported_asset = register_imported_asset(
                session,
                workspace_slug=workspace_slug,
                original_filename=original_filename,
                content=base64.b64decode(content_b64),
                mime_type=mime_type,
                preferred_asset_id=asset_id,
            )
            imported_assets[asset_id or imported_asset.id] = imported_asset.id
        except Exception as exc:
            skipped += 1
            errors.append(f"Asset bundle item failed: {exc}")

    def _resolve_asset_ids(value: Any) -> list[str]:
        if value is None or value == "":
            return []
        raw_ids = value if isinstance(value, list) else [item.strip() for item in str(value).split(",") if item.strip()]
        resolved: list[str] = []
        for raw_id in raw_ids:
            candidate = imported_assets.get(str(raw_id), str(raw_id))
            if not session.get(Asset, candidate):
                raise ValidationAPIError("Referenced asset id was not found.", details={"asset_id": candidate})
            resolved.append(candidate)
        return resolved

    for index, row in enumerate(rows, start=1):
        try:
            card_id_value = row.get("card_id")
            title = str(row.get("title") or "").strip() or "Imported Card"
            summary = str(row.get("summary") or "").strip()
            status = str(row.get("status") or "draft").strip() or "draft"
            dynamic_fields: dict[str, Any] = {}
            gallery_asset_ids: list[str] = []
            attachment_asset_ids: list[str] = []
            cover_asset_id: str | None = None
            for source_key, target_slug in matched_columns.items():
                if target_slug in {"card_id", "title", "summary", "status"}:
                    continue
                if target_slug == "gallery_asset_ids":
                    gallery_asset_ids = _resolve_asset_ids(row.get(source_key))
                    continue
                if target_slug == "attachment_asset_ids":
                    attachment_asset_ids = _resolve_asset_ids(row.get(source_key))
                    continue
                if target_slug == "cover_asset_id":
                    resolved_cover = _resolve_asset_ids(row.get(source_key))
                    cover_asset_id = resolved_cover[0] if resolved_cover else None
                    continue
                field = fields.get(target_slug)
                if field is None:
                    continue
                dynamic_fields[target_slug] = _coerce_import_value(field, row.get(source_key))

            if card_id_value not in {None, ""}:
                try:
                    card_id = int(card_id_value)
                    existing = get_card(session, card_id)
                    update_card(
                        session,
                        existing.id,
                        CardUpdate(
                            title=title,
                            summary=summary,
                            status=status,
                            schema_id=definition.slug,
                            dynamic_fields=dynamic_fields,
                        ),
                    )
                    for asset_id in gallery_asset_ids:
                        attach_existing_asset(
                            session,
                            card_id=existing.id,
                            asset_id=asset_id,
                            role="gallery",
                            set_as_cover=cover_asset_id == asset_id,
                        )
                    for asset_id in attachment_asset_ids:
                        attach_existing_asset(
                            session,
                            card_id=existing.id,
                            asset_id=asset_id,
                            role="attachment",
                        )
                    updated += 1
                    continue
                except Exception:
                    pass

            created_card = create_card(
                session,
                CardCreate(
                    title=title,
                    summary=summary,
                    status=status,
                    schema_id=definition.slug,
                    taxonomy_term_ids=[],
                ),
            )
            update_card(
                session,
                created_card.id,
                CardUpdate(
                    dynamic_fields=dynamic_fields,
                    summary=summary,
                    status=status,
                    schema_id=definition.slug,
                ),
            )
            for asset_id in gallery_asset_ids:
                attach_existing_asset(
                    session,
                    card_id=created_card.id,
                    asset_id=asset_id,
                    role="gallery",
                    set_as_cover=cover_asset_id == asset_id,
                )
            for asset_id in attachment_asset_ids:
                attach_existing_asset(
                    session,
                    card_id=created_card.id,
                    asset_id=asset_id,
                    role="attachment",
                )
            created += 1
        except Exception as exc:
            skipped += 1
            errors.append(f"Row {index}: {exc}")

    rebuild_index(session)
    session.commit()
    return {
        "card_type_slug": definition.slug,
        "format": import_format.lower(),
        "rows_created": created,
        "rows_updated": updated,
        "rows_skipped": skipped,
        "errors": errors,
        "missing_columns": preview["missing_columns"],
        "unknown_columns": preview["unknown_columns"],
    }


def export_card_type_table(session: Session, card_type_slug: str, *, export_format: str, q: str = "") -> dict[str, Any]:
    table = get_card_type_table(session, card_type_slug, q=q)
    rows = [
        {
            "card_id": row["card_id"],
            "title": row["title"],
            "summary": row["summary"],
            "status": row["status"],
            **row["values"],
        }
        for row in table["rows"]
    ]
    export_format = export_format.lower()
    if export_format == "json":
        return {
            "card_type_slug": table["card_type"]["slug"],
            "format": "json",
            "filename": f"{table['card_type']['slug']}-table.json",
            "content_json": rows,
            "content_text": "",
            "content_base64": "",
        }
    headers = list(rows[0].keys()) if rows else ["card_id", "title", "summary", "status"]
    if export_format == "csv":
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
        return {
            "card_type_slug": table["card_type"]["slug"],
            "format": "csv",
            "filename": f"{table['card_type']['slug']}-table.csv",
            "content_json": [],
            "content_text": buffer.getvalue(),
            "content_base64": "",
        }
    if export_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Card Type Table"
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        buffer = io.BytesIO()
        workbook.save(buffer)
        return {
            "card_type_slug": table["card_type"]["slug"],
            "format": "xlsx",
            "filename": f"{table['card_type']['slug']}-table.xlsx",
            "content_json": [],
            "content_text": "",
            "content_base64": base64.b64encode(buffer.getvalue()).decode("ascii"),
        }
    raise ValidationAPIError("Unsupported export format.", details={"supported": ["json", "csv", "xlsx"]})
