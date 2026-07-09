from __future__ import annotations

import importlib
import json
import os
import shutil
import sqlite3
import zipfile
import base64
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from fastapi.testclient import TestClient
from sqlalchemy.exc import OperationalError

TEST_DATA_DIR = Path(__file__).resolve().parents[2] / ".test-data"
if TEST_DATA_DIR.exists():
    shutil.rmtree(TEST_DATA_DIR)
os.environ["CODEX_DATA_DIR"] = str(TEST_DATA_DIR)

client = TestClient(importlib.import_module("app.main").app)


def _json_data(response):
    payload = response.json()
    assert "data" in payload
    return payload["data"]


def _workspace_slug() -> str:
    workspaces = _json_data(client.get("/api/workspaces"))
    assert workspaces
    return workspaces[0]["slug"]


def _png_bytes() -> bytes:
    return (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
        b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDAT\x08\xd7c\xf8\xff\xff?"
        b"\x00\x05\xfe\x02\xfeA\xd9\xa2\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
    )


def test_health_and_app_info() -> None:
    health = client.get("/api/health")
    assert health.status_code == 200
    assert _json_data(health)["status"] == "ok"

    app_info = client.get("/api/app-info")
    assert app_info.status_code == 200
    payload = _json_data(app_info)
    assert payload["project_name"] == "Personal Codex Builder API"
    assert payload["workspace_count"] >= 1


def test_workspaces_bootstrap() -> None:
    response = client.get("/api/workspaces")
    assert response.status_code == 200
    workspaces = _json_data(response)
    assert workspaces
    slug = workspaces[0]["slug"]
    assert (TEST_DATA_DIR / "app_index.sqlite").exists()
    workspace_root = TEST_DATA_DIR / "workspaces" / slug
    assert (workspace_root / "workspace.sqlite").exists()
    assert (workspace_root / "workspace_manifest.json").exists()
    assert (workspace_root / "workspace.json").exists()
    assert (workspace_root / "files").exists()
    assert (workspace_root / "backups").exists()
    assert (workspace_root / "exports").exists()


def test_workspace_reorder_persists() -> None:
    first = _json_data(
        client.post(
            "/api/workspaces",
            json={"name": "Alpha Vault", "description": "", "theme": "fantasy"},
        )
    )
    second = _json_data(
        client.post(
            "/api/workspaces",
            json={"name": "Beta Vault", "description": "", "theme": "classic"},
        )
    )

    reorder_response = client.post(
        "/api/workspaces/reorder",
        json={"ordered_slugs": [second["slug"], first["slug"]]},
    )
    assert reorder_response.status_code == 200

    workspaces = _json_data(client.get("/api/workspaces"))
    alpha_index = next(index for index, item in enumerate(workspaces) if item["slug"] == first["slug"])
    beta_index = next(index for index, item in enumerate(workspaces) if item["slug"] == second["slug"])
    assert beta_index < alpha_index


def test_create_card_and_search() -> None:
    slug = _workspace_slug()
    schema_response = client.put(
        f"/api/workspaces/{slug}/schemas/creature",
        json={
            "id": "creature",
            "label": "Creature",
            "description": "Creatures and beings",
            "icon": "C",
            "field_order": ["homeland", "threat_level"],
            "fields": [
                {
                    "field_id": "homeland",
                    "label": "Homeland",
                    "kind": "text",
                    "show_in_card": True,
                    "show_in_list": True,
                    "show_in_filters": True,
                },
                {
                    "field_id": "threat_level",
                    "label": "Threat Level",
                    "kind": "number",
                    "show_in_card": True,
                    "show_in_list": True,
                },
            ],
        },
    )
    assert schema_response.status_code == 200

    create_response = client.post(f"/api/workspaces/{slug}/cards", json={"title": "Astral Beacon"})
    assert create_response.status_code == 200
    card = _json_data(create_response)
    card_id = card["id"]
    update_response = client.patch(
        f"/api/workspaces/{slug}/cards/{card_id}",
        json={
            "schema_id": "creature",
            "summary": "A bright guide for night sailors.",
            "body_json": {
                "type": "doc",
                "content": [
                    {
                        "type": "paragraph",
                        "content": [{"type": "text", "text": "The beacon shines above the harbor."}],
                    }
                ],
            },
            "dynamic_fields": {"homeland": "Lunar Harbor", "threat_level": 3},
        },
    )
    assert update_response.status_code == 200
    search_response = client.get(f"/api/workspaces/{slug}/cards?q=beacon")
    assert search_response.status_code == 200
    assert any(item["id"] == card_id for item in _json_data(search_response)["items"])

    card_types_response = client.get(f"/api/workspaces/{slug}/card-types")
    assert card_types_response.status_code == 200
    card_types = _json_data(card_types_response)
    assert any(item["slug"] == "creature" for item in card_types)

    table_response = client.get(f"/api/workspaces/{slug}/card-types/creature/table")
    assert table_response.status_code == 200
    table_payload = _json_data(table_response)
    assert table_payload["card_type"]["slug"] == "creature"
    assert any(column["field_slug"] == "homeland" for column in table_payload["columns"])
    assert any(row["card_id"] == card_id for row in table_payload["rows"])

    filtered_table = client.get(f"/api/workspaces/{slug}/card-types/creature/table?q=lunar")
    assert filtered_table.status_code == 200
    assert _json_data(filtered_table)["total"] >= 1

    structure_export = client.get(f"/api/workspaces/{slug}/card-types/creature/structure-export?format=csv")
    assert structure_export.status_code == 200
    exported_structure = _json_data(structure_export)
    assert exported_structure["filename"].endswith(".csv")
    assert "card_type_slug" in exported_structure["content_text"]

    structure_export_xlsx = client.get(f"/api/workspaces/{slug}/card-types/creature/structure-export?format=xlsx")
    assert structure_export_xlsx.status_code == 200
    structure_xlsx_payload = _json_data(structure_export_xlsx)
    workbook = load_workbook(BytesIO(base64.b64decode(structure_xlsx_payload["content_base64"])))
    assert workbook.active.max_row >= 2

    table_export_xlsx = client.get(f"/api/workspaces/{slug}/card-types/creature/table-export?format=xlsx")
    assert table_export_xlsx.status_code == 200
    table_xlsx_payload = _json_data(table_export_xlsx)
    table_workbook = load_workbook(BytesIO(base64.b64decode(table_xlsx_payload["content_base64"])))
    assert table_workbook.active.max_row >= 2

    preview_response = client.post(
      f"/api/workspaces/{slug}/card-types/creature/import-preview",
      json={
          "format": "csv",
          "content_text": "title,summary,status,homeland,threat_level\nSky Kraken,Storm hunter,active,Maelstrom,5\n",
      },
    )
    assert preview_response.status_code == 200
    preview_payload = _json_data(preview_response)
    assert preview_payload["row_count"] == 1
    assert preview_payload["missing_columns"] == []
    assert preview_payload["matched_columns"]["homeland"] == "homeland"

    import_response = client.post(
      f"/api/workspaces/{slug}/card-types/creature/import-apply",
      json={
          "format": "json",
          "content_text": '[{"title":"Moon Witch","summary":"Guides tides","status":"active","homeland":"Silver Reef","threat_level":4}]',
      },
    )
    assert import_response.status_code == 200
    import_payload = _json_data(import_response)
    assert import_payload["rows_created"] == 1
    assert import_payload["rows_skipped"] == 0

    imported_table = client.get(f"/api/workspaces/{slug}/card-types/creature/table?q=silver")
    assert imported_table.status_code == 200
    assert _json_data(imported_table)["total"] >= 1

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["title", "summary", "status", "homeland", "threat_level"])
    sheet.append(["Star Eel", "Deep current scout", "active", "Azure Trench", 2])
    xlsx_buffer = BytesIO()
    workbook.save(xlsx_buffer)
    xlsx_b64 = base64.b64encode(xlsx_buffer.getvalue()).decode("ascii")

    xlsx_preview = client.post(
        f"/api/workspaces/{slug}/card-types/creature/import-preview",
        json={"format": "xlsx", "content_base64": xlsx_b64, "filename": "creature.xlsx"},
    )
    assert xlsx_preview.status_code == 200
    assert _json_data(xlsx_preview)["row_count"] == 1

    xlsx_import = client.post(
        f"/api/workspaces/{slug}/card-types/creature/import-apply",
        json={"format": "xlsx", "content_base64": xlsx_b64, "filename": "creature.xlsx"},
    )
    assert xlsx_import.status_code == 200
    assert _json_data(xlsx_import)["rows_created"] == 1

    asset_card = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Asset Host"}))
    second_card = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Asset Reuse"}))
    upload_response = client.post(
        f"/api/workspaces/{slug}/cards/{asset_card['id']}/assets/gallery",
        files={"upload": ("map.png", BytesIO(_png_bytes()), "image/png")},
    )
    assert upload_response.status_code == 200
    uploaded_card_detail = _json_data(upload_response)
    uploaded_asset = uploaded_card_detail["gallery"][-1]
    asset_library_after_upload = _json_data(client.get(f"/api/workspaces/{slug}/assets"))
    uploaded_library_asset = next(
        item
        for item in asset_library_after_upload["items"]
        if item["relative_path"] == uploaded_asset["stored_path"]
    )

    attach_response = client.post(
        f"/api/workspaces/{slug}/assets/{uploaded_library_asset['id']}/attach",
        json={"card_id": second_card["id"], "role": "gallery", "set_as_cover": True},
    )
    assert attach_response.status_code == 200

    second_card_detail = _json_data(client.get(f"/api/workspaces/{slug}/cards/{second_card['id']}"))
    assert second_card_detail["gallery"]
    assert second_card_detail["cover_asset_id"] == uploaded_library_asset["id"]

    cover_update_response = client.patch(
        f"/api/workspaces/{slug}/cards/{asset_card['id']}",
        json={"cover_asset_id": uploaded_library_asset["id"]},
    )
    assert cover_update_response.status_code == 200
    assert _json_data(cover_update_response)["cover_asset_id"] == uploaded_library_asset["id"]

    invalid_cover_update = client.patch(
        f"/api/workspaces/{slug}/cards/{asset_card['id']}",
        json={"cover_asset_id": "missing-asset"},
    )
    assert invalid_cover_update.status_code == 422

    delete_used_asset = client.delete(f"/api/workspaces/{slug}/assets/{uploaded_library_asset['id']}")
    assert delete_used_asset.status_code == 422

    orphan_upload = client.post(
        f"/api/workspaces/{slug}/cards/{asset_card['id']}/assets/attachment",
        files={"upload": ("note.txt", BytesIO(b'hello world'), "text/plain")},
    )
    assert orphan_upload.status_code == 200
    orphan_library = _json_data(client.get(f"/api/workspaces/{slug}/assets?q=note.txt"))
    orphan_asset_id = orphan_library["items"][0]["id"]
    remove_from_card = client.delete(
        f"/api/workspaces/{slug}/cards/assets/{orphan_asset_id}?card_id={asset_card['id']}"
    )
    assert remove_from_card.status_code == 200
    after_removal_library = _json_data(client.get(f"/api/workspaces/{slug}/assets?q=note.txt"))
    assert after_removal_library["items"] == []

    asset_library_response = client.get(f"/api/workspaces/{slug}/assets")
    assert asset_library_response.status_code == 200
    asset_library = _json_data(asset_library_response)
    assert any(item["id"] == uploaded_library_asset["id"] for item in asset_library["items"])
    assert all(item["original_filename"] != "note.txt" for item in asset_library["items"])


def test_slug_uniqueness_validation() -> None:
    slug = _workspace_slug()
    first = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "First"}))
    second = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Second"}))
    response = client.patch(
        f"/api/workspaces/{slug}/cards/{second['id']}",
        json={"slug": first["slug"]},
    )
    assert response.status_code == 409
    assert response.json()["error"]["code"] == "CONFLICT"


def test_card_delete_defaults_to_soft_delete_and_hides_card() -> None:
    slug = _workspace_slug()
    client.put(
        f"/api/workspaces/{slug}/schemas/character",
        json={
            "id": "character",
            "label": "Character",
            "description": "",
            "icon": "C",
            "field_order": ["origin"],
            "fields": [
                {
                    "field_id": "origin",
                    "label": "Origin",
                    "kind": "text",
                    "show_in_card": True,
                    "show_in_list": True,
                }
            ],
        },
    )
    card = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Soft Delete", "schema_id": "character"}))
    delete_response = client.delete(f"/api/workspaces/{slug}/cards/{card['id']}")
    assert delete_response.status_code == 200
    assert _json_data(delete_response)["message"] == "Card archived."

    detail_response = client.get(f"/api/workspaces/{slug}/cards/{card['id']}")
    assert detail_response.status_code == 404

    listed_cards = _json_data(client.get(f"/api/workspaces/{slug}/cards?q=soft delete"))["items"]
    assert all(item["id"] != card["id"] for item in listed_cards)

    table_rows = _json_data(client.get(f"/api/workspaces/{slug}/card-types/character/table"))["rows"]
    assert all(row["card_id"] != card["id"] for row in table_rows)

    workspace_db = TEST_DATA_DIR / "workspaces" / slug / "workspace.sqlite"
    with sqlite3.connect(workspace_db) as connection:
        registry_row = connection.execute(
            "SELECT status, deleted_at FROM cards_registry WHERE legacy_card_id = ?",
            (card["id"],),
        ).fetchone()
        card_row = connection.execute("SELECT status FROM cards WHERE id = ?", (card["id"],)).fetchone()
        search_row = connection.execute(
            "SELECT COUNT(*) FROM card_search_index WHERE card_id = ?",
            (str(card["id"]),),
        ).fetchone()

    assert registry_row is not None
    assert registry_row[0] == "archived"
    assert registry_row[1]
    assert card_row == ("archived",)
    assert search_row == (0,)


def test_card_type_field_changes_are_safe_and_fts_search_has_fallback(monkeypatch) -> None:
    created_workspace = _json_data(
        client.post(
            "/api/workspaces",
            json={"name": "Evolution Lab", "description": "", "theme": "fantasy"},
        )
    )
    slug = created_workspace["slug"]
    schema_payload = {
        "id": "artifact",
        "label": "Artifact",
        "description": "",
        "icon": "A",
        "field_order": ["origin", "power_level"],
        "fields": [
            {
                "field_id": "origin",
                "label": "Origin",
                "kind": "text",
                "show_in_card": True,
                "show_in_list": True,
            },
            {
                "field_id": "power_level",
                "label": "Power Level",
                "kind": "number",
                "show_in_card": True,
                "show_in_list": True,
            },
        ],
    }
    first_schema = client.put(f"/api/workspaces/{slug}/schemas/artifact", json=schema_payload)
    assert first_schema.status_code == 200

    card = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Fallback Beacon"}))
    updated_card = client.patch(
        f"/api/workspaces/{slug}/cards/{card['id']}",
        json={
            "schema_id": "artifact",
            "summary": "Ancient relay",
            "dynamic_fields": {"origin": "Old Harbor", "power_level": 7},
        },
    )
    assert updated_card.status_code == 200

    evolved_schema = client.put(
        f"/api/workspaces/{slug}/schemas/artifact",
        json={
            "id": "artifact",
            "label": "Artifact",
            "description": "Evolved safely",
            "icon": "A",
            "field_order": ["origin_realm"],
            "fields": [
                {
                    "field_id": "origin_realm",
                    "label": "Origin Realm",
                    "kind": "text",
                    "show_in_card": True,
                    "show_in_list": True,
                }
            ],
        },
    )
    assert evolved_schema.status_code == 200

    card_types = _json_data(client.get(f"/api/workspaces/{slug}/card-types"))
    artifact_type = next(item for item in card_types if item["slug"] == "artifact")
    assert artifact_type["is_active"] is True
    assert [field["field_slug"] for field in artifact_type["fields"]] == ["origin-realm"]

    table_payload = _json_data(client.get(f"/api/workspaces/{slug}/card-types/artifact/table"))
    assert table_payload["total"] == 1
    assert [column["field_slug"] for column in table_payload["columns"]] == ["origin-realm"]
    assert table_payload["rows"][0]["card_id"] == card["id"]

    workspace_db = TEST_DATA_DIR / "workspaces" / slug / "workspace.sqlite"
    with sqlite3.connect(workspace_db) as connection:
        field_rows = connection.execute(
            "SELECT field_slug, is_active FROM card_type_fields WHERE card_type_slug = 'artifact' ORDER BY id"
        ).fetchall()

    assert ("origin", 0) in field_rows
    assert ("power-level", 0) in field_rows
    assert ("origin-realm", 1) in field_rows

    import app.services.search_service as search_service

    def _raise_fts_failure(session, query):
        raise OperationalError("SELECT card_id FROM card_search_index", {"query": query}, Exception("fts unavailable"))

    monkeypatch.setattr(search_service, "_fts_matching_ids", _raise_fts_failure)
    fallback_search = client.get(f"/api/workspaces/{slug}/cards?q=fallback")
    assert fallback_search.status_code == 200
    fallback_items = _json_data(fallback_search)["items"]
    assert any(item["id"] == card["id"] for item in fallback_items)


def test_table_view_row_editing_and_workspace_data_export() -> None:
    created_workspace = _json_data(
        client.post(
            "/api/workspaces",
            json={"name": "Table Ops", "description": "", "theme": "fantasy"},
        )
    )
    slug = created_workspace["slug"]
    schema_response = client.put(
        f"/api/workspaces/{slug}/schemas/relic",
        json={
            "id": "relic",
            "label": "Relic",
            "description": "",
            "icon": "R",
            "field_order": ["rarity"],
            "fields": [
                {
                    "field_id": "rarity",
                    "label": "Rarity",
                    "kind": "text",
                    "show_in_card": True,
                    "show_in_list": True,
                }
            ],
        },
    )
    assert schema_response.status_code == 200

    create_row = client.post(
        f"/api/workspaces/{slug}/card-types/relic/rows",
        json={"title": "Sun Relic", "summary": "Warm glow", "status": "active", "values": {"rarity": "rare"}},
    )
    assert create_row.status_code == 200
    created_table = _json_data(create_row)
    created_row = next(row for row in created_table["rows"] if row["title"] == "Sun Relic")
    assert created_row["values"]["rarity"] == "rare"

    update_row = client.patch(
        f"/api/workspaces/{slug}/card-types/relic/rows/{created_row['card_id']}",
        json={"title": "Moon Relic", "summary": "Cool glow", "status": "draft", "values": {"rarity": "legendary"}},
    )
    assert update_row.status_code == 200
    updated_table = _json_data(update_row)
    updated_row = next(row for row in updated_table["rows"] if row["card_id"] == created_row["card_id"])
    assert updated_row["title"] == "Moon Relic"
    assert updated_row["values"]["rarity"] == "legendary"

    filtered_table = client.get(
        f"/api/workspaces/{slug}/card-types/relic/table?sort_by=title&sort_dir=desc&status=draft"
    )
    assert filtered_table.status_code == 200
    filtered_payload = _json_data(filtered_table)
    assert filtered_payload["sort_by"] == "title"
    assert filtered_payload["sort_dir"] == "desc"
    assert filtered_payload["status"] == "draft"
    assert filtered_payload["rows"][0]["card_id"] == created_row["card_id"]

    data_export = client.get(
        f"/api/workspaces/{slug}/data-export?format=json&card_ids={created_row['card_id']}&include_asset_ids=true"
    )
    assert data_export.status_code == 200
    exported_payload = _json_data(data_export)
    assert exported_payload["scope"] == "selected_cards"
    assert exported_payload["include_asset_ids"] is True
    assert exported_payload["row_count"] == 1
    assert exported_payload["content_json"][0]["title"] == "Moon Relic"

    archive_row = client.delete(f"/api/workspaces/{slug}/card-types/relic/rows/{created_row['card_id']}")
    assert archive_row.status_code == 200
    assert _json_data(archive_row)["deleted"] is True

    after_archive = client.get(f"/api/workspaces/{slug}/card-types/relic/table")
    assert after_archive.status_code == 200
    assert _json_data(after_archive)["rows"] == []


def test_card_type_import_supports_existing_and_packaged_assets() -> None:
    created_workspace = _json_data(
        client.post(
            "/api/workspaces",
            json={"name": "Asset Import Lab", "description": "", "theme": "fantasy"},
        )
    )
    slug = created_workspace["slug"]
    schema_response = client.put(
        f"/api/workspaces/{slug}/schemas/relic",
        json={
            "id": "relic",
            "label": "Relic",
            "description": "",
            "icon": "R",
            "field_order": ["rarity"],
            "fields": [
                {
                    "field_id": "rarity",
                    "label": "Rarity",
                    "kind": "text",
                    "show_in_card": True,
                    "show_in_list": True,
                }
            ],
        },
    )
    assert schema_response.status_code == 200

    host_card = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Host Card"}))
    upload_response = client.post(
        f"/api/workspaces/{slug}/cards/{host_card['id']}/assets/attachment",
        files={"upload": ("shared-note.txt", BytesIO(b"shared note"), "text/plain")},
    )
    assert upload_response.status_code == 200
    asset_library = _json_data(client.get(f"/api/workspaces/{slug}/assets?q=shared-note.txt"))
    existing_asset_id = asset_library["items"][0]["id"]

    bundle_asset_id = "img-bundle1"
    import_payload = {
        "rows": [
            {
                "title": "Imported Relic",
                "summary": "Came from bundle",
                "status": "active",
                "rarity": "epic",
                "attachment_asset_ids": [existing_asset_id],
                "gallery_asset_ids": [bundle_asset_id],
                "cover_asset_id": bundle_asset_id,
            }
        ],
        "asset_files": [
            {
                "asset_id": bundle_asset_id,
                "filename": "bundle-map.png",
                "mime_type": "image/png",
                "content_base64": base64.b64encode(_png_bytes()).decode("ascii"),
            }
        ],
    }

    preview_response = client.post(
        f"/api/workspaces/{slug}/card-types/relic/import-preview",
        json={"format": "json", "content_text": json.dumps(import_payload)},
    )
    assert preview_response.status_code == 200
    preview_payload = _json_data(preview_response)
    assert preview_payload["matched_columns"]["attachment_asset_ids"] == "attachment_asset_ids"
    assert preview_payload["matched_columns"]["gallery_asset_ids"] == "gallery_asset_ids"
    assert preview_payload["matched_columns"]["cover_asset_id"] == "cover_asset_id"

    import_response = client.post(
        f"/api/workspaces/{slug}/card-types/relic/import-apply",
        json={"format": "json", "content_text": json.dumps(import_payload)},
    )
    assert import_response.status_code == 200
    import_result = _json_data(import_response)
    assert import_result["rows_created"] == 1
    assert import_result["rows_skipped"] == 0

    imported_table = _json_data(client.get(f"/api/workspaces/{slug}/card-types/relic/table?q=imported"))
    imported_row = imported_table["rows"][0]
    imported_card = _json_data(client.get(f"/api/workspaces/{slug}/cards/{imported_row['card_id']}"))
    assert any(asset["id"] == existing_asset_id for asset in imported_card["attachments"])
    assert any(asset["id"] == bundle_asset_id for asset in imported_card["gallery"])
    assert imported_card["cover_asset_id"] == bundle_asset_id


def test_health_repairs_and_asset_usage_details() -> None:
    created_workspace = _json_data(
        client.post(
            "/api/workspaces",
            json={"name": "Repair Lab", "description": "", "theme": "fantasy"},
        )
    )
    slug = created_workspace["slug"]
    left = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Repair Left"}))
    right = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Repair Right"}))

    upload_response = client.post(
        f"/api/workspaces/{slug}/cards/{left['id']}/assets/gallery",
        files={"upload": ("repair-map.png", BytesIO(_png_bytes()), "image/png")},
    )
    assert upload_response.status_code == 200
    _json_data(upload_response)

    asset_library = _json_data(client.get(f"/api/workspaces/{slug}/assets?q=repair-map"))
    usage_roles = {usage["asset_role"] for usage in asset_library["items"][0]["usages"]}
    assert "gallery" in usage_roles
    assert "cover" in usage_roles

    workspace_root = TEST_DATA_DIR / "workspaces" / slug
    db_path = workspace_root / "workspace.sqlite"
    orphaned_file = workspace_root / "assets" / "other" / "orphaned-note.txt"
    orphaned_file.write_text("orphan", encoding="utf-8")

    broken = sqlite3.connect(db_path)
    try:
        broken.execute("PRAGMA foreign_keys = OFF")
        broken.execute("DELETE FROM card_search_index WHERE card_id = ?", (str(left["id"]),))
        broken.execute(
            "INSERT INTO card_relations (source_card_id, target_card_id, relation_type, note, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            (left["id"], 999999, "one-to-one", "broken", "2026-07-09T00:00:00+00:00", "2026-07-09T00:00:00+00:00"),
        )
        broken.execute("UPDATE cards SET cover_asset_id = ? WHERE id = ?", ("missing-cover", right["id"]))
        broken.commit()
    finally:
        broken.close()

    health_before = _json_data(client.get(f"/api/workspaces/{slug}/health"))
    assert any(check["key"] == "search_index_sync" and check["status"] != "ok" for check in health_before["checks"])
    assert any(check["key"] == "relation_link_integrity" and check["status"] != "ok" for check in health_before["checks"])

    asset_health_before = _json_data(client.get(f"/api/workspaces/{slug}/asset-health"))
    assert "missing-cover" in asset_health_before["broken_cover_asset_ids"]
    assert orphaned_file.relative_to(TEST_DATA_DIR / "workspaces").as_posix() in asset_health_before["orphaned_files"]

    rebuild_search = client.post(
        f"/api/workspaces/{slug}/health/repair",
        json={"action": "rebuild_search_index"},
    )
    assert rebuild_search.status_code == 200
    assert _json_data(rebuild_search)["repaired_count"] == 1

    repair_relations = client.post(
        f"/api/workspaces/{slug}/health/repair",
        json={"action": "remove_broken_relation_links"},
    )
    assert repair_relations.status_code == 200
    assert _json_data(repair_relations)["repaired_count"] >= 1

    repair_covers = client.post(
        f"/api/workspaces/{slug}/asset-health/repair",
        json={"action": "remove_broken_cover_references"},
    )
    assert repair_covers.status_code == 200
    assert _json_data(repair_covers)["repaired_count"] == 1

    repair_orphaned_files = client.post(
        f"/api/workspaces/{slug}/asset-health/repair",
        json={"action": "delete_orphaned_asset_files"},
    )
    assert repair_orphaned_files.status_code == 200
    assert _json_data(repair_orphaned_files)["repaired_count"] == 1

    health_after = _json_data(client.get(f"/api/workspaces/{slug}/health"))
    assert any(check["key"] == "search_index_sync" and check["status"] == "ok" for check in health_after["checks"])
    assert any(check["key"] == "relation_link_integrity" and check["status"] == "ok" for check in health_after["checks"])

    asset_health_after = _json_data(client.get(f"/api/workspaces/{slug}/asset-health"))
    assert asset_health_after["broken_cover_asset_ids"] == []
    assert asset_health_after["orphaned_files"] == []


def test_notebook_asset_usage_blocks_deletion() -> None:
    slug = _workspace_slug()
    host = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Notebook Host"}))

    upload_response = client.post(
        f"/api/workspaces/{slug}/cards/{host['id']}/assets/attachment",
        files={"upload": ("notebook-reference.txt", BytesIO(b"notebook asset"), "text/plain")},
    )
    assert upload_response.status_code == 200

    asset_library = _json_data(client.get(f"/api/workspaces/{slug}/assets?q=notebook-reference"))
    asset_id = asset_library["items"][0]["id"]

    notebook_update = client.patch(
        f"/api/workspaces/{slug}/notebook",
        json={
            "body_json": {"type": "doc", "content": [{"type": "paragraph"}]},
            "body_text": "",
            "items": [
                {
                    "id": "note-rich",
                    "type": "rich_text",
                    "title": "Notebook",
                    "sort_order": 0,
                    "body_json": {
                        "type": "doc",
                        "content": [{"type": "paragraph", "content": [{"type": "text", "text": "Top note"}]}],
                    },
                    "body_text": "Top note",
                },
                {
                    "id": "note-file",
                    "type": "file",
                    "title": "Attached file",
                    "sort_order": 1,
                    "asset_id": asset_id,
                    "note": "Keep this nearby",
                },
            ],
        },
    )
    assert notebook_update.status_code == 200

    asset_library_after = _json_data(client.get(f"/api/workspaces/{slug}/assets?q=notebook-reference"))
    notebook_usage = next(
        usage for usage in asset_library_after["items"][0]["usages"] if usage["usage_type"] == "notebook"
    )
    assert notebook_usage["asset_role"] == "file"

    card_unlink = client.delete(
        f"/api/workspaces/{slug}/cards/assets/{asset_id}?card_id={host['id']}"
    )
    assert card_unlink.status_code == 200

    blocked_delete = client.delete(f"/api/workspaces/{slug}/assets/{asset_id}")
    assert blocked_delete.status_code == 422

    notebook_clear = client.patch(
        f"/api/workspaces/{slug}/notebook",
        json={
            "body_json": {"type": "doc", "content": [{"type": "paragraph"}]},
            "body_text": "",
            "items": [
                {
                    "id": "note-rich",
                    "type": "rich_text",
                    "title": "Notebook",
                    "sort_order": 0,
                    "body_json": {"type": "doc", "content": [{"type": "paragraph"}]},
                    "body_text": "",
                }
            ],
        },
    )
    assert notebook_clear.status_code == 200

    delete_after_clear = client.delete(f"/api/workspaces/{slug}/assets/{asset_id}")
    assert delete_after_clear.status_code == 200


def test_sources_relations_backup_restore_export_import_and_archive() -> None:
    slug = _workspace_slug()
    left = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Left Card"}))
    right = _json_data(client.post(f"/api/workspaces/{slug}/cards", json={"title": "Right Card"}))
    source_response = client.post(
        f"/api/workspaces/{slug}/cards/{left['id']}/sources",
        json={"title": "Field Notes", "note": "Stored at the harbor", "source_type": "note"},
    )
    assert source_response.status_code == 200
    source = _json_data(source_response)

    source_asset_upload = client.post(
        f"/api/workspaces/{slug}/cards/sources/{source['id']}/assets",
        files={"upload": ("source-note.txt", BytesIO(b"source attachment"), "text/plain")},
    )
    assert source_asset_upload.status_code == 200
    source_asset_payload = _json_data(source_asset_upload)
    assert source_asset_payload["sources"][0]["assets"]
    source_asset_id = source_asset_payload["sources"][0]["assets"][0]["id"]

    source_asset_delete = client.delete(
        f"/api/workspaces/{slug}/cards/sources/assets/{source_asset_id}?source_id={source['id']}"
    )
    assert source_asset_delete.status_code == 200
    assert _json_data(source_asset_delete)["sources"][0]["assets"] == []

    library_asset_upload = client.post(
        f"/api/workspaces/{slug}/cards/{left['id']}/assets/attachment",
        files={"upload": ("shared-source.txt", BytesIO(b"shared source"), "text/plain")},
    )
    assert library_asset_upload.status_code == 200
    library_assets = _json_data(client.get(f"/api/workspaces/{slug}/assets?q=shared-source.txt"))
    shared_asset_id = library_assets["items"][0]["id"]
    attach_existing_source_asset = client.post(
        f"/api/workspaces/{slug}/cards/sources/{source['id']}/assets/{shared_asset_id}"
    )
    assert attach_existing_source_asset.status_code == 200
    attached_payload = _json_data(attach_existing_source_asset)
    assert any(asset["id"] == shared_asset_id for asset in attached_payload["sources"][0]["assets"])

    relation_response = client.post(
        f"/api/workspaces/{slug}/cards/{left['id']}/relations",
        json={"target_card_id": right["id"], "note": "Mentions"},
    )
    assert relation_response.status_code == 200
    payload = _json_data(relation_response)
    assert payload["sources"]
    assert payload["relations"][0]["target_card_id"] == right["id"]

    search_shared_source = client.get(f"/api/workspaces/{slug}/cards?q=shared-source")
    assert search_shared_source.status_code == 200
    assert any(item["id"] == left["id"] for item in _json_data(search_shared_source)["items"])

    backup_response = client.post(f"/api/workspaces/{slug}/backup", json={"reason": "manual"})
    assert backup_response.status_code == 200
    backup = _json_data(backup_response)
    assert backup["filename"].endswith(".sqlite")

    mutate_response = client.patch(
        f"/api/workspaces/{slug}/cards/{left['id']}",
        json={"summary": "Changed after backup"},
    )
    assert mutate_response.status_code == 200
    restore_response = client.post(
        f"/api/workspaces/{slug}/restore",
        json={"filename": backup["filename"]},
    )
    assert restore_response.status_code == 200
    restored = _json_data(restore_response)
    assert restored["safety_backup"]["filename"].endswith(".sqlite")
    card_after_restore = _json_data(client.get(f"/api/workspaces/{slug}/cards/{left['id']}"))
    assert card_after_restore["summary"] != "Changed after backup"

    health_response = client.get(f"/api/workspaces/{slug}/health")
    assert health_response.status_code == 200
    health = _json_data(health_response)
    assert health["integrity_ok"] is True
    assert health["backup_count"] >= 2
    assert health["issue_count"] >= 0
    assert any(check["key"] == "sqlite_integrity" for check in health["checks"])
    assert "database" in health["categories"]

    asset_health_response = client.get(f"/api/workspaces/{slug}/asset-health")
    assert asset_health_response.status_code == 200
    asset_health = _json_data(asset_health_response)
    assert "missing_asset_files" in asset_health
    assert "checks" in asset_health

    notebook_response = client.get(f"/api/workspaces/{slug}/notebook")
    assert notebook_response.status_code == 200
    notebook = _json_data(notebook_response)
    assert "body_json" in notebook
    assert "items" in notebook

    notebook_update = client.patch(
        f"/api/workspaces/{slug}/notebook",
        json={
            "body_json": {
                "type": "doc",
                "content": [{"type": "paragraph", "content": [{"type": "text", "text": "Workspace scratchpad"}]}],
            },
            "body_text": "",
            "items": [
                {
                    "id": "note-main",
                    "type": "rich_text",
                    "title": "Notebook",
                    "sort_order": 0,
                    "body_json": {
                        "type": "doc",
                        "content": [{"type": "paragraph", "content": [{"type": "text", "text": "Workspace scratchpad"}]}],
                    },
                    "body_text": "Workspace scratchpad",
                }
            ],
        },
    )
    assert notebook_update.status_code == 200
    updated_notebook = _json_data(notebook_update)
    assert updated_notebook["body_text"] == "Workspace scratchpad"
    assert updated_notebook["items"][0]["id"] == "note-main"

    export_response = client.post(f"/api/workspaces/{slug}/export")
    assert export_response.status_code == 200
    exported = _json_data(export_response)
    export_path = TEST_DATA_DIR / exported["path"]
    assert export_path.exists()
    with zipfile.ZipFile(export_path) as archive:
        members = set(archive.namelist())
    assert "workspace.sqlite" in members
    assert "workspace_manifest.json" in members
    assert "workspace.json" in members
    assert any(member.startswith("assets/") for member in members)

    with export_path.open("rb") as archive_handle:
        import_response = client.post(
            "/api/workspaces/import",
            files={"upload": ("archive.workspace.zip", archive_handle, "application/zip")},
            data={"name": "Imported Atlas"},
        )
    assert import_response.status_code == 200
    imported = _json_data(import_response)
    imported_slug = imported["slug"]
    assert imported_slug != slug
    imported_assets = _json_data(client.get(f"/api/workspaces/{imported_slug}/assets"))
    assert any(item["original_filename"] == "shared-source.txt" for item in imported_assets["items"])
    imported_asset = next(item for item in imported_assets["items"] if item["original_filename"] == "shared-source.txt")
    imported_asset_path = TEST_DATA_DIR / "workspaces" / imported_asset["relative_path"]
    assert imported_asset_path.exists()

    archive_response = client.post(f"/api/workspaces/{imported_slug}/archive")
    assert archive_response.status_code == 200
    archived = _json_data(archive_response)
    assert archived["archived"] is True

    active_workspaces = _json_data(client.get("/api/workspaces"))
    assert all(workspace["slug"] != imported_slug for workspace in active_workspaces)
    all_workspaces = _json_data(client.get("/api/workspaces?include_archived=true"))
    assert any(workspace["slug"] == imported_slug and workspace["archived"] for workspace in all_workspaces)

    unarchive_response = client.post(f"/api/workspaces/{imported_slug}/unarchive")
    assert unarchive_response.status_code == 200
    unarchived = _json_data(unarchive_response)
    assert unarchived["archived"] is False


def test_import_accepts_legacy_workspace_metadata_only() -> None:
    slug = _workspace_slug()
    export_response = client.post(f"/api/workspaces/{slug}/export")
    assert export_response.status_code == 200
    exported = _json_data(export_response)
    export_path = TEST_DATA_DIR / exported["path"]
    legacy_only_path = export_path.with_name("legacy-only.workspace.zip")

    with zipfile.ZipFile(export_path) as source_archive, zipfile.ZipFile(legacy_only_path, "w") as target_archive:
        for member in source_archive.infolist():
            if member.filename == "workspace_manifest.json":
                continue
            target_archive.writestr(member, source_archive.read(member.filename))

    with legacy_only_path.open("rb") as archive_handle:
        import_response = client.post(
            "/api/workspaces/import",
            files={"upload": ("legacy-only.workspace.zip", archive_handle, "application/zip")},
            data={"name": "Legacy Import"},
        )

    assert import_response.status_code == 200
    imported = _json_data(import_response)
    imported_root = TEST_DATA_DIR / "workspaces" / imported["slug"]
    assert (imported_root / "workspace_manifest.json").exists()
    assert (imported_root / "workspace.json").exists()
